from __future__ import annotations

import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import List, Sequence

from filelock import FileLock, Timeout
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from src.core.config import get_settings

logger = logging.getLogger(__name__)

EXPECTED_HEADERS: Sequence[str] = (
    "timestamp",
    "user_id",
    "username",
    "chat_id",
    "message_id",
    "text",
)

PLAVKA_HEADERS: Sequence[str] = (
    "id_plavka",
    "Учетный_номер",
    "Плавка_дата",
    "Номер_плавки",
    "Номер_кластера",
    "Старший_смены_плавки",
    "Первый_участник_смены_плавки",
    "Второй_участник_смены_плавки",
    "Третий_участник_смены_плавки",
    "Четвертый_участник_смены_плавки",
    "Наименование_отливки",
    "Тип_эксперемента",
    "Сектор_A_опоки",
    "Сектор_B_опоки",
    "Сектор_C_опоки",
    "Сектор_D_опоки",
    "Плавка_время_прогрева_ковша_A",
    "Плавка_время_перемещения_A",
    "Плавка_время_заливки_A",
    "Плавка_температура_заливки_A",
    "Плавка_время_прогрева_ковша_B",
    "Плавка_время_перемещения_B",
    "Плавка_время_заливки_B",
    "Плавка_температура_заливки_B",
    "Плавка_время_прогрева_ковша_C",
    "Плавка_время_перемещения_C",
    "Плавка_время_заливки_C",
    "Плавка_температура_заливки_C",
    "Плавка_время_прогрева_ковша_D",
    "Плавка_время_перемещения_D",
    "Плавка_время_заливки_D",
    "Плавка_температура_заливки_D",
    "Комментарий",
    "Плавка_время_заливки",
    "id",
)

LOCK_TIMEOUT = 15  # seconds


class ExcelServiceError(Exception):
    """Base class for Excel service errors."""


class ExcelValidationError(ExcelServiceError):
    """Raised when the Excel sheet does not match the expected structure."""


def _get_lock(path: Path) -> FileLock:
    return FileLock(f"{path}.lock", timeout=LOCK_TIMEOUT)


def _prepare_workbook(path: Path) -> None:
    if not path.exists():
        mode = _detect_workbook_mode(path)
        logger.info("Excel file not found. Creating a new workbook at %s with mode=%s", path, mode)
        workbook = Workbook()
        worksheet = workbook.active
        
        if mode == "plavka":
            worksheet.title = "Records"
            worksheet.append(list(PLAVKA_HEADERS))
        else:
            worksheet.title = "Journal"
            worksheet.append(list(EXPECTED_HEADERS))
        
        workbook.save(path)
        workbook.close()
        return

    try:
        workbook = load_workbook(path)
    except InvalidFileException as exc:
        raise ExcelValidationError(
            "Не удалось открыть plavka.xlsx. Проверьте, что файл не поврежден и используется формат XLSX."
        ) from exc

    worksheet = workbook.active
    mode = _detect_workbook_mode(path)
    
    if mode == "plavka":
        header_values = [worksheet.cell(row=1, column=index + 1).value for index in range(min(len(PLAVKA_HEADERS), worksheet.max_column))]
        
        if all(value in (None, "") for value in header_values) and worksheet.max_row == 1:
            logger.info("Excel file found without headers. Writing plavka headers to %s", path)
            for column, header in enumerate(PLAVKA_HEADERS, start=1):
                worksheet.cell(row=1, column=column, value=header)
            workbook.save(path)
            workbook.close()
            return
        
        workbook.close()
    else:
        header_values = [worksheet.cell(row=1, column=index + 1).value for index in range(len(EXPECTED_HEADERS))]

        if all(value in (None, "") for value in header_values) and worksheet.max_row == 1:
            logger.info("Excel file found without headers. Writing default headers to %s", path)
            for column, header in enumerate(EXPECTED_HEADERS, start=1):
                worksheet.cell(row=1, column=column, value=header)
            workbook.save(path)
            workbook.close()
            return

        if list(header_values) != list(EXPECTED_HEADERS):
            workbook.close()
            raise ExcelValidationError(
                "Структура листа plavka.xlsx не соответствует ожидаемой. "
                "Проверьте заголовки: timestamp, user_id, username, chat_id, message_id, text."
            )

        workbook.close()


def ensure_workbook_ready() -> None:
    settings = get_settings()
    xlsx_path = settings.xlsx_path
    lock = _get_lock(xlsx_path)

    try:
        with lock:
            _prepare_workbook(xlsx_path)
    except Timeout as exc:
        raise ExcelServiceError(
            "Файл plavka.xlsx сейчас используется. Попробуйте повторить попытку позже."
        ) from exc



def append_message_row(*, user_id: int, username: str | None, chat_id: int, message_id: int, text: str) -> None:
    settings = get_settings()
    xlsx_path = settings.xlsx_path
    lock = _get_lock(xlsx_path)

    try:
        with lock:
            _prepare_workbook(xlsx_path)

            try:
                workbook = load_workbook(xlsx_path)
            except InvalidFileException as exc:
                raise ExcelValidationError(
                    "Не удалось открыть plavka.xlsx для записи. Проверьте структуру файла."
                ) from exc

            worksheet = workbook.active
            timestamp = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
            worksheet.append(
                [
                    timestamp,
                    user_id,
                    username or "",
                    chat_id,
                    message_id,
                    text,
                ]
            )
            workbook.save(xlsx_path)
            workbook.close()
            logger.info(
                "Добавлена запись в журнал: user_id=%s, chat_id=%s, message_id=%s",
                user_id,
                chat_id,
                message_id,
            )
    except Timeout as exc:
        raise ExcelServiceError(
            "Файл plavka.xlsx сейчас используется. Попробуйте повторить попытку позже."
        ) from exc


def get_last_rows(limit: int) -> List[List[str | int | None]]:
    if limit <= 0:
        return []

    settings = get_settings()
    xlsx_path = settings.xlsx_path
    lock = _get_lock(xlsx_path)

    try:
        with lock:
            _prepare_workbook(xlsx_path)

            try:
                workbook = load_workbook(xlsx_path, read_only=True)
            except InvalidFileException as exc:
                raise ExcelValidationError(
                    "Не удалось прочитать plavka.xlsx. Проверьте структуру файла."
                ) from exc

            worksheet = workbook.active
            rows = [list(row) for row in worksheet.iter_rows(min_row=2, values_only=True)]
            workbook.close()
    except Timeout as exc:
        raise ExcelServiceError(
            "Файл plavka.xlsx сейчас используется. Попробуйте повторить попытку позже."
        ) from exc

    if not rows:
        return []

    return rows[-limit:]


def _detect_workbook_mode(path: Path) -> str:
    if not path.exists():
        return "plavka"
    
    try:
        workbook = load_workbook(path, read_only=True)
        worksheet = workbook.active
        first_row = [worksheet.cell(row=1, column=i+1).value for i in range(min(10, worksheet.max_column))]
        workbook.close()
        
        if first_row[0] == "id_plavka" or "Учетный_номер" in first_row:
            return "plavka"
        elif first_row[0] == "timestamp":
            return "journal"
        else:
            return "plavka"
    except Exception:
        return "plavka"


def append_plavka_rows(rows: List[List]) -> int:
    settings = get_settings()
    xlsx_path = settings.xlsx_path
    lock = _get_lock(xlsx_path)
    
    mode = _detect_workbook_mode(xlsx_path)
    
    if mode != "plavka" and xlsx_path.exists():
        raise ExcelValidationError(
            "Файл plavka.xlsx имеет неправильную структуру. Ожидалась структура для записи плавок."
        )

    try:
        with lock:
            if not xlsx_path.exists():
                logger.info("Creating new plavka workbook at %s", xlsx_path)
                workbook = Workbook()
                worksheet = workbook.active
                worksheet.title = "Records"
                worksheet.append(list(PLAVKA_HEADERS))
                workbook.save(xlsx_path)
                workbook.close()
            
            try:
                workbook = load_workbook(xlsx_path)
            except InvalidFileException as exc:
                raise ExcelValidationError(
                    "Не удалось открыть plavka.xlsx для записи плавок. Проверьте структуру файла."
                ) from exc
            
            worksheet = workbook.active
            header_values = [worksheet.cell(row=1, column=i+1).value for i in range(len(PLAVKA_HEADERS))]
            
            if list(header_values) != list(PLAVKA_HEADERS):
                workbook.close()
                raise ExcelValidationError(
                    f"Структура листа plavka.xlsx не соответствует ожидаемой для плавок. "
                    f"Ожидалось: {len(PLAVKA_HEADERS)} столбцов, найдено: {len(header_values)}"
                )
            
            rows_added = 0
            for row in rows:
                worksheet.append(row)
                rows_added += 1
            
            workbook.save(xlsx_path)
            workbook.close()
            logger.info("Добавлено %d плавок в журнал", rows_added)
            return rows_added
    except Timeout as exc:
        raise ExcelServiceError(
            "Файл plavka.xlsx сейчас используется. Попробуйте повторить попытку позже."
        ) from exc
