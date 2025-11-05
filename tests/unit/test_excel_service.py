"""Unit tests for Excel service."""

from __future__ import annotations

import tempfile
from pathlib import Path
from unittest.mock import patch, MagicMock
from datetime import datetime, timezone

import pytest
from openpyxl import Workbook
from filelock import Timeout

from src.bot.services.excel import (
    ExcelServiceError,
    ExcelValidationError,
    append_message_row,
    get_last_rows,
    ensure_workbook_ready,
    _prepare_workbook,
    _get_lock,
    EXPECTED_HEADERS,
)


class TestExcelService:
    """Test cases for Excel service."""

    @pytest.fixture
    def temp_excel_file(self) -> Path:
        """Create a temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            return Path(tmp.name)

    @pytest.fixture
    def mock_settings(self, temp_excel_file: Path):
        """Mock settings with temporary file path."""
        settings = MagicMock()
        settings.xlsx_path = temp_excel_file
        return settings

    def test_get_lock_creates_lock_file(self, temp_excel_file: Path):
        """Test that _get_lock creates a proper lock file path."""
        lock = _get_lock(temp_excel_file)
        assert str(lock.lock_file) == f"{temp_excel_file}.lock"

    def test_prepare_workbook_creates_new_file(self, temp_excel_file: Path):
        """Test _prepare_workbook creates new file with headers."""
        assert not temp_excel_file.exists()
        
        _prepare_workbook(temp_excel_file)
        
        assert temp_excel_file.exists()
        
        # Verify headers
        from openpyxl import load_workbook
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        headers = [ws.cell(row=1, column=i+1).value for i in range(len(EXPECTED_HEADERS))]
        assert headers == list(EXPECTED_HEADERS)
        wb.close()

    def test_prepare_workbook_handles_existing_file_with_headers(self, temp_excel_file: Path):
        """Test _prepare_workbook with existing file with correct headers."""
        # Create file with headers
        wb = Workbook()
        ws = wb.active
        ws.append(list(EXPECTED_HEADERS))
        ws.append(["2024-01-01T12:00:00", 123, "user", 456, 789, "test"])
        wb.save(temp_excel_file)
        wb.close()
        
        # Should not raise any exceptions
        _prepare_workbook(temp_excel_file)
        
        # Verify file still has correct structure
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        headers = [ws.cell(row=1, column=i+1).value for i in range(len(EXPECTED_HEADERS))]
        assert headers == list(EXPECTED_HEADERS)
        assert ws.max_row == 2  # Header + 1 data row
        wb.close()

    def test_prepare_workbook_adds_headers_to_empty_file(self, temp_excel_file: Path):
        """Test _prepare_workbook adds headers to existing empty file."""
        # Create empty workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Journal"
        wb.save(temp_excel_file)
        wb.close()
        
        _prepare_workbook(temp_excel_file)
        
        # Verify headers were added
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        headers = [ws.cell(row=1, column=i+1).value for i in range(len(EXPECTED_HEADERS))]
        assert headers == list(EXPECTED_HEADERS)
        wb.close()

    def test_prepare_workbook_raises_validation_error_for_wrong_headers(self, temp_excel_file: Path):
        """Test _prepare_workbook raises error for wrong headers."""
        # Create file with wrong headers
        wb = Workbook()
        ws = wb.active
        ws.append(["wrong", "headers", "here"])
        wb.save(temp_excel_file)
        wb.close()
        
        with pytest.raises(ExcelValidationError, match="Структура листа plavka.xlsx не соответствует ожидаемой"):
            _prepare_workbook(temp_excel_file)

    @patch('src.bot.services.excel.get_settings')
    def test_ensure_workbook_ready_success(self, mock_get_settings, temp_excel_file: Path):
        """Test ensure_workbook_ready creates/validates workbook successfully."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        
        ensure_workbook_ready()
        
        assert temp_excel_file.exists()

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_ensure_workbook_ready_timeout_error(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test ensure_workbook_ready handles timeout error."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_lock.__enter__.side_effect = Timeout("Lock timeout")
        mock_get_lock.return_value = mock_lock
        
        with pytest.raises(ExcelServiceError, match="Файл plavka.xlsx сейчас используется"):
            ensure_workbook_ready()

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_append_message_row_success(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test append_message_row adds row successfully."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        # Create initial file with headers
        _prepare_workbook(temp_excel_file)
        
        append_message_row(
            user_id=123,
            username="testuser",
            chat_id=456,
            message_id=789,
            text="test message"
        )
        
        # Verify row was added
        from openpyxl import load_workbook
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        assert ws.max_row == 2  # Header + 1 data row
        
        last_row = [ws.cell(row=2, column=i+1).value for i in range(len(EXPECTED_HEADERS))]
        assert last_row[1] == 123  # user_id
        assert last_row[2] == "testuser"  # username
        assert last_row[3] == 456  # chat_id
        assert last_row[4] == 789  # message_id
        assert last_row[5] == "test message"  # text
        assert last_row[0] is not None  # timestamp
        
        wb.close()

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_append_message_row_with_none_username(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test append_message_row handles None username."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        _prepare_workbook(temp_excel_file)
        
        append_message_row(
            user_id=123,
            username=None,
            chat_id=456,
            message_id=789,
            text="test message"
        )
        
        from openpyxl import load_workbook
        wb = load_workbook(temp_excel_file)
        ws = wb.active
        last_row = [ws.cell(row=2, column=i+1).value for i in range(len(EXPECTED_HEADERS))]
        assert last_row[2] == ""  # None username should become empty string
        wb.close()

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_append_message_row_timeout_error(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test append_message_row handles timeout error."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_lock.__enter__.side_effect = Timeout("Lock timeout")
        mock_get_lock.return_value = mock_lock
        
        with pytest.raises(ExcelServiceError, match="Файл plavka.xlsx сейчас используется"):
            append_message_row(
                user_id=123,
                username="testuser",
                chat_id=456,
                message_id=789,
                text="test message"
            )

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_append_message_row_validation_error(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test append_message_row handles validation error."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        # Create file with wrong headers to trigger validation error
        wb = Workbook()
        ws = wb.active
        ws.append(["wrong", "headers"])
        wb.save(temp_excel_file)
        wb.close()
        
        with pytest.raises(ExcelValidationError, match="Структура листа plavka.xlsx не соответствует ожидаемой"):
            append_message_row(
                user_id=123,
                username="testuser",
                chat_id=456,
                message_id=789,
                text="test message"
            )

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_get_last_rows_empty_file(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test get_last_rows with empty file."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        _prepare_workbook(temp_excel_file)
        
        rows = get_last_rows(5)
        assert rows == []

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_get_last_rows_with_data(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test get_last_rows with data."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        # Create file with headers and data
        wb = Workbook()
        ws = wb.active
        ws.append(list(EXPECTED_HEADERS))
        
        # Add 5 rows of test data
        for i in range(1, 6):
            ws.append([
                f"2024-01-0{i}T12:00:00",
                i * 100,
                f"user{i}",
                i * 1000,
                i * 10000,
                f"message {i}"
            ])
        
        wb.save(temp_excel_file)
        wb.close()
        
        rows = get_last_rows(3)
        assert len(rows) == 3
        
        # Should return last 3 rows in order
        assert rows[0][1] == 300  # user_id from row 3
        assert rows[1][1] == 400  # user_id from row 4
        assert rows[2][1] == 500  # user_id from row 5

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_get_last_rows_limit_zero(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test get_last_rows with limit=0."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        rows = get_last_rows(0)
        assert rows == []

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_get_last_rows_negative_limit(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test get_last_rows with negative limit."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        rows = get_last_rows(-5)
        assert rows == []

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_get_last_rows_limit_greater_than_data(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test get_last_rows with limit greater than available data."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_get_lock.return_value = mock_lock
        
        # Create file with only 2 rows of data
        wb = Workbook()
        ws = wb.active
        ws.append(list(EXPECTED_HEADERS))
        ws.append(["2024-01-01T12:00:00", 123, "user1", 456, 789, "message1"])
        ws.append(["2024-01-02T12:00:00", 124, "user2", 457, 790, "message2"])
        wb.save(temp_excel_file)
        wb.close()
        
        rows = get_last_rows(10)  # Request more than available
        assert len(rows) == 2  # Should only return available rows

    @patch('src.bot.services.excel.get_settings')
    @patch('src.bot.services.excel._get_lock')
    def test_get_last_rows_timeout_error(self, mock_get_lock, mock_get_settings, temp_excel_file: Path):
        """Test get_last_rows handles timeout error."""
        mock_get_settings.return_value = MagicMock(xlsx_path=temp_excel_file)
        mock_lock = MagicMock()
        mock_lock.__enter__.side_effect = Timeout("Lock timeout")
        mock_get_lock.return_value = mock_lock
        
        with pytest.raises(ExcelServiceError, match="Файл plavka.xlsx сейчас используется"):
            get_last_rows(5)