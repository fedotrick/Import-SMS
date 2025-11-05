"""Integration tests for Excel service with real file operations."""

from __future__ import annotations

import tempfile
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed
from unittest.mock import patch
import threading
import time

import pytest

from src.bot.services.excel import (
    append_message_row,
    get_last_rows,
    ensure_workbook_ready,
    ExcelServiceError,
)


class TestExcelIntegration:
    """Integration tests for Excel service."""

    @pytest.fixture
    def temp_dir(self) -> Path:
        """Create temporary directory for integration tests."""
        with tempfile.TemporaryDirectory() as tmp_dir:
            yield Path(tmp_dir)

    @pytest.fixture
    def excel_path(self, temp_dir: Path) -> Path:
        """Create Excel file path in temp directory."""
        return temp_dir / "plavka.xlsx"

    @patch('src.bot.services.excel.get_settings')
    def test_full_workflow_initialization_to_data_retrieval(self, mock_get_settings, excel_path: Path):
        """Test complete workflow from initialization to data retrieval."""
        mock_get_settings.return_value.xlsx_path = excel_path
        
        # 1. Initialize workbook
        ensure_workbook_ready()
        assert excel_path.exists()
        
        # 2. Add some records
        test_data = [
            (123, "user1", 456, 789, "First message"),
            (124, "user2", 457, 790, "Second message"),
            (125, "user3", 458, 791, "Third message"),
        ]
        
        for user_id, username, chat_id, message_id, text in test_data:
            append_message_row(
                user_id=user_id,
                username=username,
                chat_id=chat_id,
                message_id=message_id,
                text=text
            )
        
        # 3. Retrieve last records
        last_rows = get_last_rows(5)
        
        assert len(last_rows) == 3
        
        # Verify data integrity
        for i, (user_id, username, chat_id, message_id, text) in enumerate(test_data):
            row = last_rows[i]
            assert row[1] == user_id  # user_id column
            assert row[2] == username  # username column
            assert row[3] == chat_id  # chat_id column
            assert row[4] == message_id  # message_id column
            assert row[5] == text  # text column
            assert row[0] is not None  # timestamp should be set

    @patch('src.bot.services.excel.get_settings')
    def test_cyrillic_directory_path_handling(self, mock_get_settings, temp_dir: Path):
        """Test handling of Cyrillic directory paths."""
        cyrillic_dir = temp_dir / "Контроль"
        cyrillic_dir.mkdir(exist_ok=True)
        excel_path = cyrillic_dir / "plavka.xlsx"
        
        mock_get_settings.return_value.xlsx_path = excel_path
        
        # Should work without issues
        ensure_workbook_ready()
        assert excel_path.exists()
        
        append_message_row(
            user_id=123,
            username="пользователь",
            chat_id=456,
            message_id=789,
            text="Тестовое сообщение на кириллице"
        )
        
        rows = get_last_rows(1)
        assert len(rows) == 1
        assert rows[0][5] == "Тестовое сообщение на кириллице"

    @patch('src.bot.services.excel.get_settings')
    def test_concurrent_access_safety(self, mock_get_settings, excel_path: Path):
        """Test that concurrent access doesn't corrupt data."""
        mock_get_settings.return_value.xlsx_path = excel_path
        
        # Initialize workbook
        ensure_workbook_ready()
        
        # Number of concurrent workers
        num_workers = 20
        records_per_worker = 5
        
        def worker_records(worker_id: int):
            """Worker function that adds records."""
            results = []
            for i in range(records_per_worker):
                try:
                    append_message_row(
                        user_id=worker_id * 1000 + i,
                        username=f"worker_{worker_id}",
                        chat_id=worker_id * 100 + i,
                        message_id=worker_id * 10 + i,
                        text=f"Message from worker {worker_id}, record {i}"
                    )
                    results.append(True)
                except Exception as e:
                    print(f"Worker {worker_id} failed: {e}")
                    results.append(False)
            return results
        
        # Run workers concurrently
        with ThreadPoolExecutor(max_workers=num_workers) as executor:
            futures = [executor.submit(worker_records, i) for i in range(num_workers)]
            
            all_results = []
            for future in as_completed(futures):
                all_results.extend(future.result())
        
        # Verify all operations succeeded
        assert all(all_results), "Some concurrent operations failed"
        
        # Verify data integrity
        all_rows = get_last_rows(num_workers * records_per_worker + 10)  # Get more than expected
        assert len(all_rows) == num_workers * records_per_worker
        
        # Check for duplicates (shouldn't be any)
        message_ids = [row[4] for row in all_rows]  # message_id column
        assert len(message_ids) == len(set(message_ids)), "Found duplicate message IDs"

    @patch('src.bot.services.excel.get_settings')
    def test_file_persistence_across_operations(self, mock_get_settings, excel_path: Path):
        """Test that data persists across multiple operations."""
        mock_get_settings.return_value.xlsx_path = excel_path
        
        # First session: add data
        ensure_workbook_ready()
        append_message_row(
            user_id=123,
            username="user1",
            chat_id=456,
            message_id=789,
            text="Session 1 message"
        )
        
        # Simulate new session - re-initialize settings
        ensure_workbook_ready()
        
        # Add more data
        append_message_row(
            user_id=124,
            username="user2",
            chat_id=457,
            message_id=790,
            text="Session 2 message"
        )
        
        # Verify both messages are present
        rows = get_last_rows(5)
        assert len(rows) == 2
        
        texts = [row[5] for row in rows]
        assert "Session 1 message" in texts
        assert "Session 2 message" in texts

    @patch('src.bot.services.excel.get_settings')
    def test_large_dataset_handling(self, mock_get_settings, excel_path: Path):
        """Test handling of larger datasets."""
        mock_get_settings.return_value.xlsx_path = excel_path
        
        ensure_workbook_ready()
        
        # Add many records
        num_records = 100
        for i in range(num_records):
            append_message_row(
                user_id=i,
                username=f"user_{i}",
                chat_id=i * 10,
                message_id=i * 100,
                text=f"Message {i}"
            )
        
        # Test pagination-like behavior
        page_size = 10
        all_retrieved = []
        
        for offset in range(0, num_records, page_size):
            # Get records from the end
            rows = get_last_rows(page_size + offset)
            if offset == 0:
                page = rows[-page_size:] if len(rows) >= page_size else rows
            else:
                page = rows[-page_size - offset:-offset] if len(rows) > offset else []
            
            all_retrieved.extend(reversed(page))  # Reverse to get chronological order
        
        # Verify we got all records
        assert len(all_retrieved) == num_records
        
        # Spot check some records
        for i in [0, 25, 50, 75, 99]:
            found = any(row[1] == i for row in all_retrieved)  # user_id column
            assert found, f"Record {i} not found"

    @patch('src.bot.services.excel.get_settings')
    def test_error_recovery_and_file_integrity(self, mock_get_settings, excel_path: Path):
        """Test error recovery and file integrity maintenance."""
        mock_get_settings.return_value.xlsx_path = excel_path
        
        # Add some initial data
        ensure_workbook_ready()
        append_message_row(
            user_id=123,
            username="user1",
            chat_id=456,
            message_id=789,
            text="Initial message"
        )
        
        # Verify initial data
        rows = get_last_rows(5)
        assert len(rows) == 1
        
        # Try to add more data
        append_message_row(
            user_id=124,
            username="user2",
            chat_id=457,
            message_id=790,
            text="Second message"
        )
        
        # Verify both messages are present and file is not corrupted
        rows = get_last_rows(5)
        assert len(rows) == 2
        
        # Verify headers are still intact
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb.active
        headers = [ws.cell(row=1, column=i+1).value for i in range(6)]
        expected_headers = ["timestamp", "user_id", "username", "chat_id", "message_id", "text"]
        assert headers == expected_headers
        wb.close()

    @patch('src.bot.services.excel.get_settings')
    def test_performance_with_multiple_operations(self, mock_get_settings, excel_path: Path):
        """Test performance with multiple rapid operations."""
        mock_get_settings.return_value.xlsx_path = excel_path
        
        ensure_workbook_ready()
        
        # Measure time for multiple operations
        start_time = time.time()
        
        num_operations = 50
        for i in range(num_operations):
            append_message_row(
                user_id=i,
                username=f"user_{i}",
                chat_id=i * 10,
                message_id=i * 100,
                text=f"Performance test message {i}"
            )
        
        append_time = time.time() - start_time
        
        # Measure read time
        start_time = time.time()
        rows = get_last_rows(num_operations)
        read_time = time.time() - start_time
        
        # Verify correctness
        assert len(rows) == num_operations
        
        # Performance should be reasonable (these are loose bounds)
        assert append_time < 10.0, f"Append operations took too long: {append_time}s"
        assert read_time < 5.0, f"Read operations took too long: {read_time}s"
        
        print(f"Performance: {num_operations} appends in {append_time:.2f}s, read in {read_time:.2f}s")