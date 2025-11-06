#!/usr/bin/env python3
"""Test concurrent writes to Excel file with simpler test."""

import sys
import time
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).parent.parent))

# Set test environment
os.environ['XLSX_PATH'] = './Контроль/plavka_test.xlsx'

from src.bot.services.excel import append_plavka_rows
from src.bot.services.parser import PlavkaRecord


def create_test_plavka(index: int) -> PlavkaRecord:
    """Create a test plavka record."""
    return PlavkaRecord(
        id_plavka=202411000 + index,
        uchetny_nomer=f"11-{index}/24",
        plavka_data=datetime.now(),
        nomer_plavki=f"11-{index}",
        nomer_klastera=None,
        starshiy_smeny="Test Worker",
        perviy_uchastnik="Worker 1",
        vtoroy_uchastnik=None,
        tretiy_uchastnik=None,
        chetvertyy_uchastnik=None,
        naimenovanie_otlivki=f"Test Item {index}",
        tip_eksperementa=None,
        sektor_a_opoki=None,
        sektor_b_opoki=None,
        sektor_c_opoki=None,
        sektor_d_opoki=None,
        plavka_vremya_progreva_kovsha_a=None,
        plavka_vremya_peremesheniya_a=None,
        plavka_vremya_zalivki_a=None,
        plavka_temperatura_zalivki_a=1520.0 + index,
        plavka_vremya_progreva_kovsha_b=None,
        plavka_vremya_peremesheniya_b=None,
        plavka_vremya_zalivki_b=None,
        plavka_temperatura_zalivki_b=None,
        plavka_vremya_progreva_kovsha_c=None,
        plavka_vremya_peremesheniya_c=None,
        plavka_vremya_zalivki_c=None,
        plavka_temperatura_zalivki_c=None,
        plavka_vremya_progreva_kovsha_d=None,
        plavka_vremya_peremesheniya_d=None,
        plavka_vremya_zalivki_d=None,
        plavka_temperatura_zalivki_d=None,
        kommentariy=f"Concurrent test {index}",
        plavka_vremya_zalivki=None,
    )


def write_worker(worker_id: int, num_records: int) -> tuple[int, bool, str]:
    """Worker function to write records concurrently."""
    try:
        start_index = worker_id * num_records
        plavki = [create_test_plavka(start_index + i) for i in range(num_records)]
        rows = [plavka.to_excel_row(start_index + i + 1) for i, plavka in enumerate(plavki)]
        
        start_time = time.time()
        rows_added = append_plavka_rows(rows)
        elapsed = time.time() - start_time
        
        return worker_id, True, f"Worker {worker_id}: wrote {rows_added} rows in {elapsed:.2f}s"
    except Exception as e:
        return worker_id, False, f"Worker {worker_id}: failed with {e}"


def test_concurrent_writes(num_workers: int = 5, records_per_worker: int = 2):
    """Test concurrent writes to Excel file."""
    print(f"\nTest: Concurrent writes with {num_workers} workers, {records_per_worker} records each")
    print("-" * 60)
    
    # Clean up test file
    test_file = Path('./Контроль/plavka_test.xlsx')
    if test_file.exists():
        test_file.unlink()
    
    start_time = time.time()
    results = []
    
    with ThreadPoolExecutor(max_workers=num_workers) as executor:
        futures = [
            executor.submit(write_worker, i, records_per_worker)
            for i in range(num_workers)
        ]
        
        for future in as_completed(futures):
            worker_id, success, message = future.result()
            results.append(success)
            status = "✓" if success else "✗"
            print(f"{status} {message}")
    
    elapsed = time.time() - start_time
    
    success_count = sum(results)
    print("-" * 60)
    print(f"Total time: {elapsed:.2f}s")
    print(f"Success rate: {success_count}/{num_workers} workers")
    
    expected_total = num_workers * records_per_worker
    print(f"Expected records: {expected_total}")
    
    # Verify file
    from openpyxl import load_workbook
    wb = load_workbook(test_file)
    ws = wb.active
    actual_rows = ws.max_row - 1  # Subtract header
    wb.close()
    print(f"Actual records in file: {actual_rows}")
    
    return success_count == num_workers and actual_rows == expected_total


def main():
    print("=" * 60)
    print("EXCEL CONCURRENT WRITE TEST (Simplified)")
    print("=" * 60)
    
    tests = [
        (lambda: test_concurrent_writes(5, 2), "5 workers, 2 records each"),
        (lambda: test_concurrent_writes(10, 1), "10 workers, 1 record each"),
    ]
    
    results = []
    for test_func, description in tests:
        print(f"\n{'=' * 60}")
        print(f"Running: {description}")
        print(f"{'=' * 60}")
        results.append(test_func())
    
    print("\n" + "=" * 60)
    print(f"Results: {sum(results)}/{len(results)} tests passed")
    print("=" * 60)
    
    # Cleanup
    test_file = Path('./Контроль/plavka_test.xlsx')
    if test_file.exists():
        test_file.unlink()
        print("\n✓ Cleaned up test file")
    
    return all(results)


if __name__ == "__main__":
    success = main()
    sys.exit(0 if success else 1)
