import pandas as pd
from openpyxl import load_workbook
import re
import os
from datetime import datetime
import logging

# Настройка логирования
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

EXCEL_FILE = 'plavka.xlsx'

def parse_import_message(text):
    """
    Парсит сообщение извлекает из него данные о плавках
    """
    # Парсим дату смены из шапки
    m = re.search(r'📅 Дата: ([0-9]{2}\.[0-9]{2}\.[0-9]{4})', text)
    plavka_date = m.group(1) if m else ''
    
    # Парсим старшего смены и участников из шапки
    m = re.search(r'👨‍💼 Старший: ([^\n]+)', text)
    starshiy = m.group(1).strip() if m else ''
    
    # Парсим участников
    uchastniki = []
    participants_match = re.search(r'👥 Участники \(\d+\):([\s\S]*?)(?=\n\n🔥 ДЕТАЛИ ПЛАВОК:|$)', text)
    if participants_match:
        participants_text = participants_match.group(1)
        # Извлекаем имена участников из списка
        participants = re.findall(r'• ([^\n]+)', participants_text)
        uchastniki = [p.strip() for p in participants]
    
    # Разбиваем на блоки по плавкам
    blocks = re.split(r'\n(?=✅ \d+\. |🔄 \d+\. )', text)
    results = []
    
    for block in blocks:
        if 'Плавка' not in block:
            continue
            
        data = {}
        # Дата смены для каждой плавки
        data['Плавка_дата'] = plavka_date
        
        # Старший смены и участники для каждой плавки
        data['Старший_смены_плавки'] = starshiy
        for i, field in enumerate(['Первый_участник_смены_плавки', 'Второй_участник_смены_плавки', 'Третий_участник_смены_плавки', 'Четвертый_участник_смены_плавки']):
            data[field] = uchastniki[i] if i < len(uchastniki) else ''
        
        # Плавка (Учетный номер)
        m = re.search(r'Плавка ([0-9]+-[0-9]+/[0-9]{2})', block)
        if m:
            data['Учетный_номер'] = m.group(1).strip()
        
        # Маршрутная карта
        m = re.search(r'📋 Маршрутная карта: (\d+)', block)
        if m:
            data['Маршрутная_карта'] = m.group(1).strip()
        
        # Кластер
        m = re.search(r'🏷️ Кластер: ([^\n]+)', block)
        if m:
            data['Номер_кластера'] = m.group(1).strip()
        
        # Наименование отливки
        m = re.search(r'🏭 Отливка: ([^\n]+)', block)
        if m:
            data['Наименование_отливки'] = m.group(1).strip()
        
        # Тип эксперимента (Литниковая система)
        m = re.search(r'⚙️ Литниковая система: ([^\n]+)', block)
        if m:
            data['Тип_эксперемента'] = m.group(1).strip()
        
        # Опоки
        m = re.search(r'📦 Опоки:\s*([^\n]+)', block)
        opoki = []
        if m:
            opoki = [x.strip().replace('Опока №', '') for x in m.group(1).split(',')]
            opoki = [str(int(o)) if o.isdigit() or (o.replace('.','',1).isdigit() and float(o).is_integer()) else o for o in opoki]
        
        # Температура
        m = re.search(r'🌡️ Температура: ([0-9]+[.,]?[0-9]*)', block)
        temp = float(m.group(1).replace(',', '.').replace('°C', '')) if m else None
        
        # Время заливки
        m = re.search(r'⏰ Время заливки: ([0-9]{2}:[0-9]{2})', block)
        time_val = m.group(1).strip() if m else ''
        
        # Комментарий
        m = re.search(r'💬 Комментарий: ([^\n]+)', block)
        comment = m.group(1).strip() if m else ''
        
        # Установка общего времени заливки
        data['Плавка_время_заливки'] = time_val
        
        # Установка комментария
        data['Комментарий'] = comment
        
        # Заполняем сектора (опоки и температуры)
        for i, sector in enumerate(['A', 'B', 'C', 'D']):
            if i < len(opoki):
                data[f'Сектор_{sector}_опоки'] = opoki[i]
                data[f'Плавка_температура_заливки_{sector}'] = temp
            else:
                data[f'Сектор_{sector}_опоки'] = ''
                data[f'Плавка_температура_заливки_{sector}'] = None
        results.append(data)
    return results


def parse_plavka_number(num):
    # num: строка вида '5-102' или '05-102'
    try:
        parts = str(num).split('-')
        if len(parts) == 2:
            nnn = parts[1].zfill(3)
        else:
            nnn = str(num).zfill(3)
        return nnn
    except Exception:
        return str(num).zfill(3)


def generate_id_plavka(date_str, num):
    """
    Генерирует id_plavka из строки даты в формате DD.MM.YYYY
    """
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        nnn = parse_plavka_number(num)
        return f"{date_obj.year}{date_obj.month:02d}{nnn}"
    except Exception as e:
        logger.error(f"Ошибка при генерации id_plavka: {e}")
        return ""


def generate_uchet_number(date_str, num):
    """
    Генерирует учетный номер из строки даты в формате DD.MM.YYYY
    """
    try:
        date_obj = datetime.strptime(date_str, '%d.%m.%Y')
        nnn = parse_plavka_number(num)
        return f"{date_obj.month:02d}-{nnn}/{str(date_obj.year)[-2:]}"
    except Exception as e:
        logger.error(f"Ошибка при генерации учетного номера: {e}")
        return ""


def add_to_excel(new_row):
    """
    Добавляет новую строку в Excel файл
    """
    try:
        # Создаем резервную копию файла, если он существует
        if os.path.exists(EXCEL_FILE):
            backup_file = f"{EXCEL_FILE}.bak"
            try:
                import shutil
                shutil.copy2(EXCEL_FILE, backup_file)
            except Exception as e:
                logger.warning(f"Не удалось создать резервную копию: {str(e)}")
        
        if not os.path.exists(EXCEL_FILE):
            # Если файл не существует, создаем новый с заголовками
            df = pd.DataFrame([new_row])
            try:
                df.to_excel(EXCEL_FILE, index=False, engine='openpyxl')
                logger.info(f"Данные сохранены в новый файл {EXCEL_FILE}")
                return True
            except PermissionError:
                logger.error("Не удалось создать файл Excel. Проверьте, не открыт ли он другой программой.")
                return False
            except Exception as e:
                logger.error(f"Ошибка при создании Excel файла: {str(e)}")
                return False
        
        # Загружаем существующий файл
        max_attempts = 3
        attempt = 0
        while attempt < max_attempts:
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
                
                # Проверяем, совпадают ли заголовки
                headers = [cell.value for cell in ws[1]]
                
                # Добавляем новую строку
                row_values = []
                for header in headers:
                    if header in new_row:
                        row_values.append(new_row[header])
                    else:
                        row_values.append(None)
                
                ws.append(row_values)
                
                # Сохраняем изменения
                wb.save(EXCEL_FILE)
                # Удаляем резервную копию после успешного сохранения
                if os.path.exists(f"{EXCEL_FILE}.bak"):
                    try:
                        os.remove(f"{EXCEL_FILE}.bak")
                    except:
                        pass
                logger.info(f"Данные также сохранены в Excel файл {EXCEL_FILE}")
                return True
                
            except PermissionError:
                attempt += 1
                if attempt < max_attempts:
                    logger.warning(f"Файл Excel занят. Попытка {attempt} из {max_attempts}...")
                    import time
                    time.sleep(1)  # Ждем 1 секунду перед повторной попыткой
                else:
                    logger.error("Не удалось сохранить в Excel. Файл занят другой программой.")
                    return False
            except Exception as e:
                logger.error(f"Ошибка при работе с Excel: {str(e)}")
                # Восстанавливаем из резервной копии при ошибке
                if os.path.exists(f"{EXCEL_FILE}.bak"):
                    try:
                        import shutil
                        shutil.copy2(f"{EXCEL_FILE}.bak", EXCEL_FILE)
                        logger.info("Восстановлена резервная копия Excel файла")
                    except:
                        pass
                return False
    except Exception as e:
        logger.error(f"Критическая ошибка при сохранении в Excel: {str(e)}")
        return False


def import_message_to_excel(message_text):
    """
    Основная функция для импорта сообщения в Excel
    """
    try:
        # Парсим сообщение
        parsed_data = parse_import_message(message_text)
        
        if not parsed_data:
            logger.warning("Не удалось распознать данные в сообщении")
            return False, "Не удалось распознать данные в сообщении"
        
        # Обрабатываем каждую плавку из сообщения
        for i, data in enumerate(parsed_data):
            # Генерируем id_plavka и учетный номер, если их нет
            if 'Учетный_номер' in data and data['Учетный_номер']:
                uchet = data['Учетный_номер']
                # Если формат учетного номера правильный, извлекаем из него компоненты для id_plavka
                if re.match(r'[0-9]{2}-[0-9]{3}/[0-9]{2}', uchet):
                    mm, nnn_yy = uchet.split('-')
                    nnn, yy = nnn_yy.split('/')  # nnn может быть разной длины
                    year = int('20' + yy)
                    month = int(mm)
                    id_plavka = f"{year}{month:02d}{nnn}"
                    номер_плавки = f"{month}-{nnn}"
                else:
                    # Если учетного номера нет, генерируем компоненты из даты
                    if data.get('Плавка_дата') and re.match(r'[0-9]{2}\.[0-9]{2}\.[0-9]{4}', data['Плавка_дата']):
                        day, month, year = map(int, data['Плавка_дата'].split('.'))
                        # Генерируем случайный номер плавки, если его нет
                        nnn = "000" # Неизвестный номер
                        id_plavka = f"{year}{month:02d}{nnn}"
                        номер_плавки = f"{month}-{nnn}"
                    else:
                        id_plavka = ""
                        номер_плавки = ""
            else:
                # Генерируем учетный номер и id_plavka, если нет Учетный_номер
                if data.get('Плавка_дата'):
                    номер_плавки = f"{data['Плавка_дата'][:2]}-000"  # Временный номер
                    id_plavka = generate_id_plavka(data['Плавка_дата'], номер_плавки)
                    uchet = generate_uchet_number(data['Плавка_дата'], номер_плавки)
                else:
                    id_plavka = ""
                    номер_плавки = ""
                    uchet = ""
            
            # Обновляем данные с правильными значениями
            data['id_plavka'] = id_plavka
            data['Учетный_номер'] = uchet
            data['Номер_плавки'] = номер_плавки
            
            # Добавляем id (автоинкремент)
            data['id'] = i + 1  # В реальном приложении можно заменить на реальный id из БД
            
            # Добавляем в Excel
            success = add_to_excel(data)
            if not success:
                logger.error(f"Не удалось сохранить плавку {i+1} в Excel")
                return False, f"Ошибка при сохранении плавки {i+1}"
        
        logger.info(f"Успешно импортировано {len(parsed_data)} плавок из сообщения")
        return True, f"Успешно импортировано {len(parsed_data)} плавок"
        
    except Exception as e:
        logger.error(f"Ошибка при импорте сообщения: {str(e)}")
        return False, f"Ошибка при импорте: {str(e)}"


# Пример использования
if __name__ == "__main__":
    # Пример сообщения для тестирования
    test_message = """
    📅 Дата: 01.11.2025
    👨‍💼 Старший: Петров
    👥 Участники (4):
    • Иванов
    • Сидоров
    • Козлов
    • Новиков

    🔥 ДЕТАЛИ ПЛАВОК:
    ✅ 1. Плавка 11-001/25
    🏷️ Кластер: 5
    🏭 Отливка: Вороток
    ⚙️ Литниковая система: Бумага
    📦 Опоки: 123, 124, 125, 126
    🌡️ Температура: 1550°C
    ⏰ Время заливки: 14:30
    💬 Комментарий: Нормальная плавка
    📋 Маршрутная карта: 12345
    """
    
    success, message = import_message_to_excel(test_message)
    print(f"Результат импорта: {success}, {message}")